import os
import argparse
from openpyxl import load_workbook, Workbook

# Create the command line arguments parser
parser = argparse.ArgumentParser(description='Process Excel files')
parser.add_argument('--folder', required=True, help='Directory where your Excel files are')
parser.add_argument('--output', required=True, help='Path to save the output Excel file')
parser.add_argument('--prefix', required=False, help='Prefix of filenames to be processed')

# Parse the command line arguments
args = parser.parse_args()

# Names to check in the first column
names = ["VERNEMURAKAMI", "ROBERTDENNIS", "JASONDLUGOS"]

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Iterate over each file in the directory
for filename in os.listdir(args.folder):
    # Process files that end with .xlsx and start with the provided prefix
    if filename.endswith(".xlsx") and (args.prefix is None or filename.startswith(args.prefix)): 
        print(f'Processing file: {filename}')
        # Construct full file path
        file_path = os.path.join(args.folder, filename)
        
        # Load the Excel file
        workbook = load_workbook(filename=file_path)

        # Iterate over each sheet in the Excel file
        for sheet in workbook.worksheets:
            # Append the first row of each file
            ws.append([cell.value for cell in sheet[1]])

            # Iterate over each row in the sheet
            for row in sheet.iter_rows(min_row=2):
                # Check the conditions in the first and second column
                if row[0].value in names and row[1].value == "Missed Meals Allowance   $":
                    # Append the row values as a list to the new sheet
                    ws.append([cell.value for cell in row])

# Save the new workbook
wb.save(args.output)

######### Second Portion - Select Engineer Records and Transpose

# Load the aggregated workbook
wb = load_workbook(filename=args.output)

# Get the active worksheet
ws = wb.active

# Create a new workbook for the result
new_wb = Workbook()
new_ws = new_wb.active

# Initialize the current row for transposed header
current_row = 1

# Initialize a dictionary to store the mapping of names to columns
name_to_column = {}

# Initialize the current column for transposed data
current_column = len(name_to_column) + 2  # We start at column 2 and add the number of unique names

# Flag for first header
first_header = True

# Iterate over each row in the worksheet
for row in ws.iter_rows():
    # Check if the cell in the first column has the value "Employee Name"
    if row[0].value == "Employee Name":
        # If this is not the first header, increment current row by 19
        if not first_header:
            current_row += 19
        else:
            first_header = False

        # Iterate over the cells in the row and add them to the new worksheet
        for i, cell in enumerate(row[:19]):  # The header row has 19 columns
            new_ws.cell(row=current_row+i, column=1, value=cell.value)
    else:
        # If it's not a header row, it's a data row
        # Get the name from the first column
        name = row[0].value

        # If this name hasn't been encountered before, map it to a new column
        if name not in name_to_column:
            name_to_column[name] = current_column
            current_column += 1

        # Transpose the data and write it to the column mapped to this name
        for i, cell in enumerate(row[:19]):  # The data row has 19 columns
            new_ws.cell(row=current_row+i, column=name_to_column[name], value=cell.value)

# After transposing all data, remove undesired rows
for row in new_ws['A']:
    if row.value in ['Wk1', 'Wk2', 'Total', 'Transaction']:
        new_ws.delete_rows(row.row)

# Save the new workbook
new_wb.save(args.output + '_transposed.xlsx')
